from pathlib import Path
import json
import openpyxl
import csv
import requests
import time

# Dictionnaire des alias de marques pour normaliser les écritures
BRAND_ALIASES = {
    "Alfa Romeo": ["Alfa Romeo", "Alfa", "Alfa-Romeo"],
    "Aston Martin": ["Aston Martin", "Aston"],
    "Audi": ["Audi"],
    "Bentley": ["Bentley"],
    "BMW": ["BMW", "Bmw"],
    "Bugatti": ["Bugatti"],
    "Cadillac": ["Cadillac"],
    "Chevrolet": ["Chevrolet", "Chevy"],
    "Chrysler": ["Chrysler"],
    "Citroën": ["Citroen", "Citroën"],
    "Cupra": ["Cupra"],
    "Dacia": ["Dacia"],
    "Daihatsu": ["Daihatsu"],
    "Dodge": ["Dodge"],
    "DS": ["DS", "DS Automobiles"],
    "Ferrari": ["Ferrari"],
    "Fiat": ["Fiat"],
    "Ford": ["Ford"],
    "Genesis": ["Genesis"],
    "Honda": ["Honda"],
    "Hummer": ["Hummer"],
    "Hyundai": ["Hyundai"],
    "Infiniti": ["Infiniti"],
    "Isuzu": ["Isuzu"],
    "Jaguar": ["Jaguar"],
    "Jeep": ["Jeep"],
    "Kia": ["Kia"],
    "Lada": ["Lada"],
    "Lamborghini": ["Lamborghini"],
    "Lancia": ["Lancia"],
    "Land Rover": ["Land Rover", "Landrover", "Land-Rover"],
    "Lexus": ["Lexus"],
    "Lotus": ["Lotus"],
    "Maserati": ["Maserati"],
    "Mazda": ["Mazda"],
    "McLaren": ["McLaren", "Mclaren"],
    "Mercedes-Benz": ["Mercedes-Benz", "Mercedes", "Mercedes Benz"],
    "MG": ["MG", "Mg"],
    "Mini": ["Mini", "MINI"],
    "Mitsubishi": ["Mitsubishi"],
    "Nissan": ["Nissan"],
    "Opel": ["Opel"],
    "Peugeot": ["Peugeot"],
    "Polestar": ["Polestar"],
    "Porsche": ["Porsche"],
    "Renault": ["Renault"],
    "Rolls-Royce": ["Rolls-Royce", "Rolls Royce", "Rollsroyce"],
    "Saab": ["Saab"],
    "Seat": ["Seat", "SEAT"],
    "Skoda": ["Skoda", "Škoda"],
    "Smart": ["Smart", "SMART"],
    "SsangYong": ["SsangYong", "Ssangyong"],
    "Subaru": ["Subaru"],
    "Suzuki": ["Suzuki"],
    "Tesla": ["Tesla"],
    "Toyota": ["Toyota"],
    "Volkswagen": ["Volkswagen", "VW"],
    "Volvo": ["Volvo"],
    "Omoda": ["Omoda"],
    "BYD": ["BYD"],
    "Lynk & Co": ["Lynk & Co", "Lynk&Co", "Lynk"],
    "Aiways": ["Aiways"],
    "Nio": ["Nio", "NIO"],
    "Rivian": ["Rivian"],
    "Lucid": ["Lucid"],
    "Fisker": ["Fisker"]
}

# Créer un mapping inverse pour rapidement normaliser une marque
ALIAS_TO_CANONICAL = {}
for canonical, aliases in BRAND_ALIASES.items():
    for alias in aliases:
        ALIAS_TO_CANONICAL[alias.lower()] = canonical

# Dictionnaire des variantes de modèles populaires avec générations
MODEL_GENERATIONS = {
    # Volkswagen
    "Golf": ["Golf", "Golf I", "Golf II", "Golf III", "Golf IV", "Golf V", "Golf VI", "Golf VII", "Golf VIII",
             "Golf 1", "Golf 2", "Golf 3", "Golf 4", "Golf 5", "Golf 6", "Golf 7", "Golf 8",
             "Golf Mk1", "Golf Mk2", "Golf Mk3", "Golf Mk4", "Golf Mk5", "Golf Mk6", "Golf Mk7", "Golf Mk8"],
    "Passat": ["Passat", "Passat B1", "Passat B2", "Passat B3", "Passat B4", "Passat B5", 
               "Passat B6", "Passat B7", "Passat B8", "Passat B9"],
    "Polo": ["Polo", "Polo I", "Polo II", "Polo III", "Polo IV", "Polo V", "Polo VI",
             "Polo 1", "Polo 2", "Polo 3", "Polo 4", "Polo 5", "Polo 6"],
    "Tiguan": ["Tiguan", "Tiguan I", "Tiguan II", "Tiguan III", "Tiguan 1", "Tiguan 2", "Tiguan 3"],
    "Touareg": ["Touareg", "Touareg I", "Touareg II", "Touareg III", "Touareg 1", "Touareg 2", "Touareg 3"],
    
    # Renault
    "Megane": ["Megane", "Mégane", "Megane I", "Megane II", "Megane III", "Megane IV", "Megane V",
               "Mégane I", "Mégane II", "Mégane III", "Mégane IV", "Mégane V",
               "Megane 1", "Megane 2", "Megane 3", "Megane 4", "Megane 5"],
    "Clio": ["Clio", "Clio I", "Clio II", "Clio III", "Clio IV", "Clio V",
             "Clio 1", "Clio 2", "Clio 3", "Clio 4", "Clio 5"],
    "Scenic": ["Scenic", "Scénic", "Scenic I", "Scenic II", "Scenic III", "Scenic IV",
               "Scénic I", "Scénic II", "Scénic III", "Scénic IV",
               "Scenic 1", "Scenic 2", "Scenic 3", "Scenic 4"],
    "Laguna": ["Laguna", "Laguna I", "Laguna II", "Laguna III", "Laguna 1", "Laguna 2", "Laguna 3"],
    
    # Peugeot
    "308": ["308", "308 I", "308 II", "308 III", "308 1", "308 2", "308 3"],
    "3008": ["3008", "3008 I", "3008 II", "3008 1", "3008 2"],
    "208": ["208", "208 I", "208 II", "208 1", "208 2"],
    "2008": ["2008", "2008 I", "2008 II", "2008 1", "2008 2"],
    
    # BMW
    "Serie 3": ["3 Series", "Serie 3", "E21", "E30", "E36", "E46", "E90", "E91", "E92", "E93", "F30", "F31", "F34", "G20", "G21"],
    "Serie 5": ["5 Series", "Serie 5", "E12", "E28", "E34", "E39", "E60", "E61", "F10", "F11", "F07", "G30", "G31"],
    "Serie 7": ["7 Series", "Serie 7", "E23", "E32", "E38", "E65", "E66", "F01", "F02", "G11", "G12"],
    "X3": ["X3", "X3 E83", "X3 F25", "X3 G01"],
    "X5": ["X5", "X5 E53", "X5 E70", "X5 F15", "X5 G05"],
    
    # Mercedes
    "Classe A": ["A-Class", "Classe A", "A-Klasse", "W168", "W169", "W176", "W177"],
    "Classe C": ["C-Class", "Classe C", "C-Klasse", "W202", "W203", "W204", "W205", "W206"],
    "Classe E": ["E-Class", "Classe E", "E-Klasse", "W210", "W211", "W212", "W213"],
    "Classe S": ["S-Class", "Classe S", "S-Klasse", "W220", "W221", "W222", "W223"],
    
    # Ford
    "Focus": ["Focus", "Focus I", "Focus II", "Focus III", "Focus IV", "Focus 1", "Focus 2", "Focus 3", "Focus 4"],
    "Fiesta": ["Fiesta", "Fiesta I", "Fiesta II", "Fiesta III", "Fiesta IV", "Fiesta V", "Fiesta VI", "Fiesta VII", "Fiesta VIII",
               "Fiesta 1", "Fiesta 2", "Fiesta 3", "Fiesta 4", "Fiesta 5", "Fiesta 6", "Fiesta 7", "Fiesta 8"],
    "Mondeo": ["Mondeo", "Mondeo I", "Mondeo II", "Mondeo III", "Mondeo IV", "Mondeo V",
               "Mondeo 1", "Mondeo 2", "Mondeo 3", "Mondeo 4", "Mondeo 5"],
    
    # Opel
    "Astra": ["Astra", "Astra F", "Astra G", "Astra H", "Astra J", "Astra K", "Astra L"],
    "Corsa": ["Corsa", "Corsa A", "Corsa B", "Corsa C", "Corsa D", "Corsa E", "Corsa F"],
    
    # Audi
    "A3": ["A3", "A3 8L", "A3 8P", "A3 8V", "A3 8Y"],
    "A4": ["A4", "A4 B5", "A4 B6", "A4 B7", "A4 B8", "A4 B9"],
    "A6": ["A6", "A6 C4", "A6 C5", "A6 C6", "A6 C7", "A6 C8"],
    "Q5": ["Q5", "Q5 8R", "Q5 FY"],
}

def ajouter_variantes_generationnelles(merged_dict: dict) -> dict:
    """
    Ajoute les variantes générationnelles pour les modèles populaires
    Ex: Si "Golf" existe, ajoute Golf I, Golf II, etc.
    """
    print("\n🔧 Ajout des variantes générationnelles...")
    
    for make_lower, data in merged_dict.items():
        modeles_ajoutés = 0
        
        # Parcourir les modèles connus avec générations
        for base_model, variantes in MODEL_GENERATIONS.items():
            # Vérifier si un des modèles correspond
            for existing_model in list(data["models"]):
                if base_model.lower() in existing_model.lower():
                    # Ajouter toutes les variantes
                    avant = len(data["models"])
                    data["models"].update(variantes)
                    après = len(data["models"])
                    modeles_ajoutés += (après - avant)
                    break
        
        if modeles_ajoutés > 0:
            print(f"  ✓ {data['make']}: +{modeles_ajoutés} variantes générationnelles")
    
    return merged_dict


def charger_wikidata_modeles(marque_nom: str, marque_wikidata_id: str = None) -> list:
    """
    Charge les modèles d'une marque depuis Wikidata
    
    Args:
        marque_nom: Nom de la marque (ex: "Volkswagen")
        marque_wikidata_id: ID Wikidata (ex: "Q246" pour VW). Si None, cherche automatiquement
    
    Returns:
        Liste des modèles trouvés
    """
    try:
        # Si pas d'ID fourni, chercher le Q-code de la marque
        if not marque_wikidata_id:
            print(f"  🔍 Recherche de l'ID Wikidata pour {marque_nom}...")
            search_url = "https://www.wikidata.org/w/api.php"
            search_params = {
                "action": "wbsearchentities",
                "search": marque_nom,
                "language": "en",
                "format": "json"
            }
            response = requests.get(search_url, params=search_params, timeout=10)
            results = response.json().get("search", [])
            
            if not results:
                print(f"  ❌ Marque '{marque_nom}' non trouvée sur Wikidata")
                return []
            
            marque_wikidata_id = results[0]["id"]
            print(f"  ✅ ID trouvé: {marque_wikidata_id}")
        
        # Requête SPARQL pour récupérer tous les modèles ET leurs générations
        sparql_url = "https://query.wikidata.org/sparql"
        query = f"""
        SELECT DISTINCT ?modelLabel WHERE {{
          ?model wdt:P176 wd:{marque_wikidata_id} .
          
          # Capturer tous les types de véhicules
          {{
            # Modèles automobiles (Q3231690)
            ?model wdt:P31 wd:Q3231690 .
          }}
          UNION
          {{
            # Générations automobiles (Q20165) - CRUCIAL pour Golf I, II, etc.
            ?model wdt:P31 wd:Q20165 .
          }}
          UNION
          {{
            # Véhicules motorisés en général
            ?model wdt:P31/wdt:P279* wd:Q1420 .
          }}
          UNION
          {{
            # Modèles de concept car
            ?model wdt:P31 wd:Q850270 .
          }}
          
          SERVICE wikibase:label {{ bd:serviceParam wikibase:language "en,fr,de" . }}
        }}
        ORDER BY ?modelLabel
        LIMIT 1000
        """
        
        params = {
            "query": query,
            "format": "json"
        }
        
        print(f"  📡 Interrogation Wikidata pour {marque_nom}...")
        response = requests.get(sparql_url, params=params, timeout=30)
        data = response.json()
        
        modeles = []
        if "results" in data and "bindings" in data["results"]:
            for binding in data["results"]["bindings"]:
                if "modelLabel" in binding:
                    model_name = binding["modelLabel"]["value"]
                    modeles.append(model_name)
        
        print(f"  ✅ {len(modeles)} modèles trouvés pour {marque_nom}")
        return modeles
        
    except requests.exceptions.Timeout:
        print("  ⏱️ Timeout lors de la requête Wikidata")
        return []
    except Exception as e:
        print(f"  ⚠️ Erreur Wikidata: {e}")
        return []

def enrichir_avec_wikidata(merged_dict: dict, marques_wikidata: dict = None) -> dict:
    """
    Enrichit les données fusionnées avec les modèles de Wikidata
    
    Args:
        merged_dict: Dictionnaire de marques/modèles fusionnés
        marques_wikidata: Dict optionnel {marque: id_wikidata}
                         Ex: {"Volkswagen": "Q246", "BMW": "Q26678"}
    
    Returns:
        Dictionnaire enrichi
    """
    if marques_wikidata is None:
        marques_wikidata = {}
    
    print("\n🌐 Enrichissement avec Wikidata...")
    
    for make_lower, data in merged_dict.items():
        marque_name = data["make"]
        modeles_avant = len(data["models"])
        
        # Chercher l'ID Wikidata si fourni
        wiki_id = marques_wikidata.get(marque_name)
        
        # Charger depuis Wikidata
        modeles_wiki = charger_wikidata_modeles(marque_name, wiki_id)
        
        if modeles_wiki:
            data["models"].update(modeles_wiki)
            modeles_apres = len(data["models"])
            ajouts = modeles_apres - modeles_avant
            print(f"  ✓ {marque_name}: +{ajouts} modèles depuis Wikidata")
            time.sleep(0.5)  # Respecter les limites de l'API
        
    return merged_dict

# IDs Wikidata pour les marques principales (optionnel, ils seront cherchés auto sinon)
WIKIDATA_MARQUES = {
    "Volkswagen": "Q246",
    "BMW": "Q26678",
    "Mercedes-Benz": "Q17305",
    "Audi": "Q23317",
    "Porsche": "Q40993",
    "Ford": "Q44294",
    "General Motors": "Q154957",
    "Toyota": "Q53327",
    "Honda": "Q9584",
    "Nissan": "Q20165",
    "Mazda": "Q215272",
    "Hyundai": "Q55931",
    "Kia": "Q35349",
    "Renault": "Q6686",
    "Peugeot": "Q6393",
    "Citroën": "Q6746",
    "Fiat": "Q27597",
    "Volvo": "Q55141",
    "Jaguar": "Q26424",
    "Land Rover": "Q34218",
    "Tesla": "Q478214",
}

def normaliser_marque(marque: str) -> str:
    """Normalise une marque en utilisant le dictionnaire d'alias"""
    if not marque:
        return ""
    marque_clean = marque.strip()
    canonical = ALIAS_TO_CANONICAL.get(marque_clean.lower())
    return canonical if canonical else marque_clean

def fusionner_datasets():
    """Fusionne EUROPEAN CARS DATASET.xlsx, autoscout24-germany-dataset.csv avec vehicle models.json"""
    
    # Chemins des fichiers
    excel_path = Path("data/references/EUROPEAN CARS DATASET.xlsx")
    csv_path = Path("data/references/autoscout24-germany-dataset.csv")
    json_path = Path("data/references/vehicle models.json")
    output_path = Path("data/references/vehicle_models_merged.json")
    
    # Vérifier l'existence des fichiers
    if not excel_path.exists():
        print(f"❌ Fichier Excel introuvable: {excel_path}")
        return
    
    if not json_path.exists():
        print(f"❌ Fichier JSON introuvable: {json_path}")
        return
    
    # Charger le JSON existant
    print("📥 Chargement du JSON existant...")
    with open(json_path, "r", encoding="utf-8") as f:
        json_data = json.load(f)
    
    # Normaliser les marques dans le JSON
    json_data_normalized = []
    for item in json_data:
        canonical = normaliser_marque(item["Make"])
        json_data_normalized.append({
            "Make": canonical,
            "Models": item["Models"]
        })
    json_data = json_data_normalized
    
    json_dict = {item["Make"].lower(): item["Models"] for item in json_data}
    print(f"✅ {len(json_dict)} marques dans le JSON (après normalisation)")
    
    # Charger le fichier Excel
    print("📥 Chargement du fichier Excel...")
    try:
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook.active
        print(f"✅ Feuille active: {worksheet.title}")
    except Exception as e:
        print(f"❌ Erreur lors de la lecture du Excel: {e}")
        return
    
    # Extraire les données du Excel
    excel_data = {}
    headers = {}
    
    # Lire les en-têtes (première ligne)
    for col_idx, cell in enumerate(worksheet[1], 1):
        if cell.value:
            headers[col_idx] = cell.value
    
    print(f"✅ Colonnes trouvées: {list(headers.values())}")
    
    # Trouver les colonnes de marque et modèle
    make_col = None
    model_col = None
    
    for col_idx, header in headers.items():
        if header and isinstance(header, str):
            header_lower = header.lower()
            if any(word in header_lower for word in ["make", "marque", "brand", "manufacturer"]):
                make_col = col_idx
            elif any(word in header_lower for word in ["model", "modèle", "name"]):
                model_col = col_idx
    
    if not make_col or not model_col:
        print("⚠️ Colonnes Marque/Modèle non trouvées")
        print(f"   Colonnes disponibles: {list(headers.values())}")
        return
    
    print(f"✅ Colonne Marque: {headers[make_col]} (col {make_col})")
    print(f"✅ Colonne Modèle: {headers[model_col]} (col {model_col})")
    
    # Extraire marques et modèles du Excel
    print("\n📊 Extraction des données Excel...")
    for row_idx in range(2, min(worksheet.max_row + 1, 12)):  # Debug: afficher les 10 premières lignes
        make = worksheet.cell(row_idx, make_col).value
        model = worksheet.cell(row_idx, model_col).value
        print(f"  Ligne {row_idx}: Make='{make}' | Model='{model}'")
    
    for row_idx in range(2, worksheet.max_row + 1):
        make = worksheet.cell(row_idx, make_col).value
        model = worksheet.cell(row_idx, model_col).value
        
        if make and model:
            make_clean = str(make).strip()
            model_clean = str(model).strip()
            
            # Normaliser la marque
            make_canonical = normaliser_marque(make_clean)
            
            # Vérifier si la marque est dans le modèle (erreur possible)
            if make_clean.lower() in model_clean.lower():
                # Enlever la marque du modèle
                model_clean = model_clean.replace(make_clean, "").strip()
                model_clean = model_clean.lstrip("-").lstrip().rstrip()
            
            make_lower = make_canonical.lower()
            
            if make_lower not in excel_data:
                excel_data[make_lower] = {"name": make_canonical, "models": []}
            
            if model_clean and model_clean not in excel_data[make_lower]["models"]:
                excel_data[make_lower]["models"].append(model_clean)
    
    print(f"✅ {len(excel_data)} marques extraites du Excel")
    
    # Charger le fichier CSV
    csv_data = {}
    if csv_path.exists():
        print("\n📥 Chargement du fichier CSV...")
        encodages = ["utf-8", "cp1252", "latin-1", "iso-8859-1"]
        
        for enc in encodages:
            try:
                with open(csv_path, "r", encoding=enc, newline="") as f:
                    reader = csv.DictReader(f)
                    headers_csv = reader.fieldnames
                    
                    print(f"✅ CSV chargé avec {enc}")
                    print(f"✅ Colonnes CSV: {headers_csv}")
                    
                    # Trouver les colonnes marque/modèle
                    make_col_csv = None
                    model_col_csv = None
                    
                    for col in headers_csv:
                        col_lower = col.lower()
                        if any(word in col_lower for word in ["make", "marque", "brand", "manufacturer"]):
                            make_col_csv = col
                        elif any(word in col_lower for word in ["model", "modèle", "name"]):
                            model_col_csv = col
                    
                    if not make_col_csv or not model_col_csv:
                        print("⚠️ Colonnes Marque/Modèle non trouvées dans le CSV")
                        break
                    
                    print(f"✅ Colonne Marque CSV: {make_col_csv}")
                    print(f"✅ Colonne Modèle CSV: {model_col_csv}")
                    
                    # Extraire les données
                    print("\n📊 Extraction des données CSV (10 premières lignes)...")
                    row_count = 0
                    for row in reader:
                        make = row.get(make_col_csv)
                        model = row.get(model_col_csv)
                        
                        if row_count < 10:
                            print(f"  Ligne {row_count+1}: Make='{make}' | Model='{model}'")
                        
                        if make and model:
                            make_clean = str(make).strip()
                            model_clean = str(model).strip()
                            
                            # Normaliser la marque
                            make_canonical = normaliser_marque(make_clean)
                            
                            # Vérifier si la marque est dans le modèle
                            if make_clean.lower() in model_clean.lower():
                                model_clean = model_clean.replace(make_clean, "").strip()
                                model_clean = model_clean.lstrip("-").lstrip().rstrip()
                            
                            make_lower = make_canonical.lower()
                            
                            if make_lower not in csv_data:
                                csv_data[make_lower] = {"name": make_canonical, "models": []}
                            
                            if model_clean and model_clean not in csv_data[make_lower]["models"]:
                                csv_data[make_lower]["models"].append(model_clean)
                        
                        row_count += 1
                    
                    print(f"✅ {len(csv_data)} marques extraites du CSV ({row_count} lignes)")
                    break
                    
            except (UnicodeDecodeError, LookupError):
                continue
            except Exception as e:
                print(f"⚠️ Erreur avec {enc}: {e}")
                continue
    else:
        print(f"\n⚠️ Fichier CSV introuvable: {csv_path}")
    
    # Fusionner les données
    print("\n🔄 Fusion des données...")
    merged_dict = {}
    
    # D'abord, ajouter toutes les données du JSON
    for item in json_data:
        make_lower = item["Make"].lower()
        merged_dict[make_lower] = {
            "make": item["Make"],
            "models": set(item["Models"])
        }
    
    # Ensuite, fusionner avec les données du Excel
    for make_lower, data in excel_data.items():
        models = data["models"]
        make_name = data["name"]
        
        if make_lower in merged_dict:
            # Marque existe: fusionner les modèles
            before_count = len(merged_dict[make_lower]["models"])
            merged_dict[make_lower]["models"].update(models)
            after_count = len(merged_dict[make_lower]["models"])
            added = after_count - before_count
            print(f"  ✓ {merged_dict[make_lower]['make']}: +{added} nouveaux modèles ({len(models)} dans Excel)")
        else:
            # Nouvelle marque: créer une entrée
            merged_dict[make_lower] = {
                "make": make_name,
                "models": set(models)
            }
            print(f"  ✓ {make_name}: {len(models)} modèles (NOUVEAU)")
    
    # Enfin, fusionner avec les données du CSV
    for make_lower, data in csv_data.items():
        models = data["models"]
        make_name = data["name"]
        
        if make_lower in merged_dict:
            # Marque existe: fusionner les modèles
            before_count = len(merged_dict[make_lower]["models"])
            merged_dict[make_lower]["models"].update(models)
            after_count = len(merged_dict[make_lower]["models"])
            added = after_count - before_count
            print(f"  ✓ {merged_dict[make_lower]['make']}: +{added} nouveaux modèles ({len(models)} dans CSV)")
        else:
            # Nouvelle marque: créer une entrée
            merged_dict[make_lower] = {
                "make": make_name,
                "models": set(models)
            }
            print(f"  ✓ {make_name}: {len(models)} modèles (NOUVEAU depuis CSV)")
    
    # Ajouter les variantes générationnelles (Golf I, II, III, etc.)
    merged_dict = ajouter_variantes_generationnelles(merged_dict)
    
    # NOUVEAU: Enrichir avec Wikidata
    merged_dict = enrichir_avec_wikidata(merged_dict, WIKIDATA_MARQUES)
    
    # Convertir en format JSON
    final_json = []
    for make_lower, data in sorted(merged_dict.items()):
        final_json.append({
            "Make": data["make"],
            "Models": sorted(list(data["models"]))
        })
    
    # Sauvegarder le JSON fusionné
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(final_json, f, ensure_ascii=False, indent=2)
    
    print("\n✅ Fusion complète!")
    print(f"✅ {len(final_json)} marques dans le JSON fusionné")
    print(f"✅ Fichier sauvegardé: {output_path}")
    
    # Statistiques
    total_models = sum(len(item["Models"]) for item in final_json)
    print(f"✅ {total_models} modèles au total")

if __name__ == "__main__":
    fusionner_datasets()
