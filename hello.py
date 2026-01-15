from pathlib import Path
import json
import openpyxl
import csv

def fusionner_datasets():
    """Fusionne EUROPEAN CARS DATASET.xlsx, autoscout24-germany-dataset.csv avec vehicle models.json"""
    
    # Chemins des fichiers
    excel_path = Path("EUROPEAN CARS DATASET.xlsx")
    csv_path = Path("autoscout24-germany-dataset.csv")
    json_path = Path("vehicle models.json")
    output_path = Path("vehicle_models_merged.json")
    
    # Vérifier l'existence des fichiers
    if not excel_path.exists():
        print(f"❌ Fichier Excel introuvable: {excel_path}")
        return
    
    if not json_path.exists():
        print(f"❌ Fichier JSON introuvable: {json_path}")
        return
    
    # Charger le JSON existant
    print("📥 Chargement du JSON existant...")
    with open(json_path, "r", encoding="utf-8") as f:
        json_data = json.load(f)
    
    json_dict = {item["Make"].lower(): item["Models"] for item in json_data}
    print(f"✅ {len(json_dict)} marques dans le JSON")
    
    # Charger le fichier Excel
    print("📥 Chargement du fichier Excel...")
    try:
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook.active
        print(f"✅ Feuille active: {worksheet.title}")
    except Exception as e:
        print(f"❌ Erreur lors de la lecture du Excel: {e}")
        return
    
    # Extraire les données du Excel
    excel_data = {}
    headers = {}
    
    # Lire les en-têtes (première ligne)
    for col_idx, cell in enumerate(worksheet[1], 1):
        if cell.value:
            headers[col_idx] = cell.value
    
    print(f"✅ Colonnes trouvées: {list(headers.values())}")
    
    # Trouver les colonnes de marque et modèle
    make_col = None
    model_col = None
    
    for col_idx, header in headers.items():
        if header and isinstance(header, str):
            header_lower = header.lower()
            if any(word in header_lower for word in ["make", "marque", "brand", "manufacturer"]):
                make_col = col_idx
            elif any(word in header_lower for word in ["model", "modèle", "name"]):
                model_col = col_idx
    
    if not make_col or not model_col:
        print(f"⚠️ Colonnes Marque/Modèle non trouvées")
        print(f"   Colonnes disponibles: {list(headers.values())}")
        return
    
    print(f"✅ Colonne Marque: {headers[make_col]} (col {make_col})")
    print(f"✅ Colonne Modèle: {headers[model_col]} (col {model_col})")
    
    # Extraire marques et modèles du Excel
    print("\n📊 Extraction des données Excel...")
    for row_idx in range(2, min(worksheet.max_row + 1, 12)):  # Debug: afficher les 10 premières lignes
        make = worksheet.cell(row_idx, make_col).value
        model = worksheet.cell(row_idx, model_col).value
        print(f"  Ligne {row_idx}: Make='{make}' | Model='{model}'")
    
    for row_idx in range(2, worksheet.max_row + 1):
        make = worksheet.cell(row_idx, make_col).value
        model = worksheet.cell(row_idx, model_col).value
        
        if make and model:
            make_clean = str(make).strip()
            model_clean = str(model).strip()
            
            # Vérifier si la marque est dans le modèle (erreur possible)
            if make_clean.lower() in model_clean.lower():
                # Enlever la marque du modèle
                model_clean = model_clean.replace(make_clean, "").strip()
                model_clean = model_clean.lstrip("-").lstrip().rstrip()
            
            make_lower = make_clean.lower()
            
            if make_lower not in excel_data:
                excel_data[make_lower] = {"name": make_clean, "models": []}
            
            if model_clean and model_clean not in excel_data[make_lower]["models"]:
                excel_data[make_lower]["models"].append(model_clean)
    
    print(f"✅ {len(excel_data)} marques extraites du Excel")
    
    # Charger le fichier CSV
    csv_data = {}
    if csv_path.exists():
        print("\n📥 Chargement du fichier CSV...")
        encodages = ["utf-8", "cp1252", "latin-1", "iso-8859-1"]
        
        for enc in encodages:
            try:
                with open(csv_path, "r", encoding=enc, newline="") as f:
                    reader = csv.DictReader(f)
                    headers_csv = reader.fieldnames
                    
                    print(f"✅ CSV chargé avec {enc}")
                    print(f"✅ Colonnes CSV: {headers_csv}")
                    
                    # Trouver les colonnes marque/modèle
                    make_col_csv = None
                    model_col_csv = None
                    
                    for col in headers_csv:
                        col_lower = col.lower()
                        if any(word in col_lower for word in ["make", "marque", "brand", "manufacturer"]):
                            make_col_csv = col
                        elif any(word in col_lower for word in ["model", "modèle", "name"]):
                            model_col_csv = col
                    
                    if not make_col_csv or not model_col_csv:
                        print(f"⚠️ Colonnes Marque/Modèle non trouvées dans le CSV")
                        break
                    
                    print(f"✅ Colonne Marque CSV: {make_col_csv}")
                    print(f"✅ Colonne Modèle CSV: {model_col_csv}")
                    
                    # Extraire les données
                    print("\n📊 Extraction des données CSV (10 premières lignes)...")
                    row_count = 0
                    for row in reader:
                        make = row.get(make_col_csv)
                        model = row.get(model_col_csv)
                        
                        if row_count < 10:
                            print(f"  Ligne {row_count+1}: Make='{make}' | Model='{model}'")
                        
                        if make and model:
                            make_clean = str(make).strip()
                            model_clean = str(model).strip()
                            
                            # Vérifier si la marque est dans le modèle
                            if make_clean.lower() in model_clean.lower():
                                model_clean = model_clean.replace(make_clean, "").strip()
                                model_clean = model_clean.lstrip("-").lstrip().rstrip()
                            
                            make_lower = make_clean.lower()
                            
                            if make_lower not in csv_data:
                                csv_data[make_lower] = {"name": make_clean, "models": []}
                            
                            if model_clean and model_clean not in csv_data[make_lower]["models"]:
                                csv_data[make_lower]["models"].append(model_clean)
                        
                        row_count += 1
                    
                    print(f"✅ {len(csv_data)} marques extraites du CSV ({row_count} lignes)")
                    break
                    
            except (UnicodeDecodeError, LookupError):
                continue
            except Exception as e:
                print(f"⚠️ Erreur avec {enc}: {e}")
                continue
    else:
        print(f"\n⚠️ Fichier CSV introuvable: {csv_path}")
    
    # Fusionner les données
    print("\n🔄 Fusion des données...")
    merged_dict = {}
    
    # D'abord, ajouter toutes les données du JSON
    for item in json_data:
        make_lower = item["Make"].lower()
        merged_dict[make_lower] = {
            "make": item["Make"],
            "models": set(item["Models"])
        }
    
    # Ensuite, fusionner avec les données du Excel
    for make_lower, data in excel_data.items():
        models = data["models"]
        make_name = data["name"]
        
        if make_lower in merged_dict:
            # Marque existe: fusionner les modèles
            before_count = len(merged_dict[make_lower]["models"])
            merged_dict[make_lower]["models"].update(models)
            after_count = len(merged_dict[make_lower]["models"])
            added = after_count - before_count
            print(f"  ✓ {merged_dict[make_lower]['make']}: +{added} nouveaux modèles ({len(models)} dans Excel)")
        else:
            # Nouvelle marque: créer une entrée
            merged_dict[make_lower] = {
                "make": make_name,
                "models": set(models)
            }
            print(f"  ✓ {make_name}: {len(models)} modèles (NOUVEAU)")
    
    # Enfin, fusionner avec les données du CSV
    for make_lower, data in csv_data.items():
        models = data["models"]
        make_name = data["name"]
        
        if make_lower in merged_dict:
            # Marque existe: fusionner les modèles
            before_count = len(merged_dict[make_lower]["models"])
            merged_dict[make_lower]["models"].update(models)
            after_count = len(merged_dict[make_lower]["models"])
            added = after_count - before_count
            print(f"  ✓ {merged_dict[make_lower]['make']}: +{added} nouveaux modèles ({len(models)} dans CSV)")
        else:
            # Nouvelle marque: créer une entrée
            merged_dict[make_lower] = {
                "make": make_name,
                "models": set(models)
            }
            print(f"  ✓ {make_name}: {len(models)} modèles (NOUVEAU depuis CSV)")
    
    # Convertir en format JSON
    final_json = []
    for make_lower, data in sorted(merged_dict.items()):
        final_json.append({
            "Make": data["make"],
            "Models": sorted(list(data["models"]))
        })
    
    # Sauvegarder le JSON fusionné
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(final_json, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ Fusion complète!")
    print(f"✅ {len(final_json)} marques dans le JSON fusionné")
    print(f"✅ Fichier sauvegardé: {output_path}")
    
    # Statistiques
    total_models = sum(len(item["Models"]) for item in final_json)
    print(f"✅ {total_models} modèles au total")

if __name__ == "__main__":
    fusionner_datasets()
